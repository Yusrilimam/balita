from flask import (
    Flask, render_template, request, redirect, url_for, 
    session, flash, jsonify, make_response, current_app, Response, send_file,
)
from collections import defaultdict, Counter
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps
import sqlite3
import numpy as np
from sklearn.preprocessing import MinMaxScaler
from models.knn import KNN
from models.lvq import LVQ
from datetime import datetime
from pathlib import Path
from sklearn.model_selection import train_test_split, cross_val_score, StratifiedKFold
from sklearn.metrics import accuracy_score, classification_report, confusion_matrix
from sklearn.base import BaseEstimator, ClassifierMixin
import os
import pdfkit
import re
import pandas as pd
from dateutil.relativedelta import relativedelta
from flask_login import login_required
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO

# =============================================
# KONFIGURASI APLIKASI
# =============================================
app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY') or 'g1z1_2025_xk32!@#something-long'
app.config['DATABASE'] = 'gizi_balita.db'


# =============================================
# FILTER TEMPLATE
# =============================================

@app.template_filter('datetime_format')
def datetime_format_filter(value):
    """Filter untuk memformat tanggal dan waktu dalam template Jinja2"""
    if value is None:
        return ""
    try:
        value = datetime.strptime(value, '%Y-%m-%d %H:%M:%S')
    except ValueError:
        try:
            value = datetime.strptime(value, '%Y-%m-%d')
        except ValueError:
            return value
    return value.strftime('%d-%m-%Y %H:%M')

# =============================================
# FUNGSI DATABASE
# =============================================

def get_db():
    """Membuat koneksi database baru"""
    conn = sqlite3.connect(app.config['DATABASE'])
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn

def init_db():
    """Inisialisasi database dengan struktur dan data awal"""
    db_path = app.config['DATABASE']
    if Path(db_path).exists():
        Path(db_path).unlink()
    conn = sqlite3.connect(db_path)
    c = conn.cursor()
    try:
        c.execute("PRAGMA foreign_keys = ON")
        
        # Buat tabel-tabel
        c.execute('''
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password TEXT NOT NULL,
            nama_lengkap TEXT NOT NULL,
            role TEXT NOT NULL CHECK (role IN ('admin', 'user')),
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
        ''')
        c.execute('''
        CREATE TABLE IF NOT EXISTS balita (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nama TEXT NOT NULL,
            tanggal_lahir DATE NOT NULL,
            jenis_kelamin TEXT NOT NULL CHECK (jenis_kelamin IN ('L', 'P')),
            nama_ortu TEXT NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
        ''')

        c.execute('''
        CREATE TABLE IF NOT EXISTS pengukuran (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            balita_id INTEGER NOT NULL,
            tanggal_ukur DATE NOT NULL,
            berat_badan REAL NOT NULL CHECK (berat_badan > 0),
            tinggi_badan REAL NOT NULL CHECK (tinggi_badan > 0),
            lingkar_lengan REAL NOT NULL CHECK (lingkar_lengan > 0),
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (balita_id) REFERENCES balita (id) ON DELETE CASCADE
        )
        ''')

        c.execute('''
        CREATE TABLE IF NOT EXISTS klasifikasi (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            pengukuran_id INTEGER UNIQUE NOT NULL,
            status_gizi TEXT NOT NULL CHECK (status_gizi IN ('normal', 'kurang', 'lebih', 'buruk')),
            tanggal_klasifikasi DATE NOT NULL,
            FOREIGN KEY (pengukuran_id) REFERENCES pengukuran (id) ON DELETE CASCADE
        )
        ''')

        c.execute('''
        CREATE TABLE IF NOT EXISTS dataset_training (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            feature1 REAL NOT NULL,
            feature2 REAL NOT NULL,
            feature3 REAL NOT NULL,
            target TEXT NOT NULL
        )
        ''')

        c.execute('''
        CREATE TABLE IF NOT EXISTS parameter_knn (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nilai_k INTEGER NOT NULL CHECK (nilai_k > 0),
            bobot_berat REAL NOT NULL CHECK (bobot_berat >= 0 AND bobot_berat <= 1),
            bobot_tinggi REAL NOT NULL CHECK (bobot_tinggi >= 0 AND bobot_tinggi <= 1),
            bobot_lila REAL NOT NULL CHECK (bobot_lila >= 0 AND bobot_lila <= 1),
            bobot_umur REAL NOT NULL CHECK (bobot_umur >= 0 AND bobot_umur <= 1),
            bobot_jk REAL NOT NULL CHECK (bobot_jk >= 0 AND bobot_jk <= 1),
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            CONSTRAINT bobot_valid CHECK (
            ABS((bobot_berat + bobot_tinggi + bobot_lila + bobot_umur + bobot_jk) - 1.0) < 0.0001
            )
        )
        ''')

        c.execute('''
        CREATE TABLE IF NOT EXISTS parameter_history (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            parameter_id INTEGER NOT NULL,
            changed_by INTEGER NOT NULL,
            nilai_k INTEGER NOT NULL,
            bobot_berat REAL NOT NULL,
            bobot_tinggi REAL NOT NULL,
            bobot_lila REAL NOT NULL,
            bobot_umur REAL NOT NULL,
            bobot_jk REAL NOT NULL,
            changed_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (parameter_id) REFERENCES parameter_knn (id),
            FOREIGN KEY (changed_by) REFERENCES users (id)
        )
        ''')

        # Insert data awal
        admin_password = os.environ.get('ADMIN_PASSWORD', 'admin')
        hashed_pw = generate_password_hash(admin_password)
        c.execute('''
            INSERT OR IGNORE INTO users (username, password, nama_lengkap, role)
            VALUES (?, ?, ?, ?)
        ''', ('admin', hashed_pw, 'Administrator', 'admin'))
        
        c.execute('''
            INSERT OR IGNORE INTO parameter_knn 
            (nilai_k, bobot_berat, bobot_tinggi, bobot_lila, bobot_umur, bobot_jk)
            VALUES (3, 0.35, 0.30, 0.15, 0.15, 0.05)
        ''')
        
        conn.commit()
        print("✅ Database berhasil diinisialisasi")
    except Exception as e:
        conn.rollback()
        print(f"❌ Error inisialisasi database: {str(e)}")
        raise
    finally:
        conn.close()

# =============================================
# DEKORATOR
# =============================================

def login_required(f):
    """Dekorator untuk memeriksa apakah pengguna telah login"""
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            flash('Silakan login terlebih dahulu', 'danger')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated

def admin_required(f):
    """Dekorator untuk memeriksa apakah pengguna adalah admin"""
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'role' not in session or session['role'] != 'admin':
            flash('Akses terbatas untuk admin', 'danger')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated

# =============================================
# ROUTE BERANDA
# =============================================

@app.route('/')
def home():
    # Koneksi ke database
    conn = sqlite3.connect('gizi_balita.db')
    cursor = conn.cursor()

    # Query data berdasarkan tahun
    cursor.execute("""
        SELECT strftime('%Y', tanggal_klasifikasi) AS tahun, status_gizi, COUNT(*) AS jumlah
        FROM klasifikasi
        GROUP BY tahun, status_gizi
        ORDER BY tahun
    """)
    rows = cursor.fetchall()

    # Format data untuk Chart.js
    data_by_year = {}
    for row in rows:
        tahun, status_gizi, jumlah = row
        if tahun not in data_by_year:
            data_by_year[tahun] = {'normal': 0, 'kurang': 0, 'lebih': 0, 'buruk': 0}
        data_by_year[tahun][status_gizi] = jumlah

    labels = list(data_by_year.keys())  # Tahun sebagai label
    normal_data = [data_by_year[tahun]['normal'] for tahun in labels]
    kurang_data = [data_by_year[tahun]['kurang'] for tahun in labels]
    lebih_data = [data_by_year[tahun]['lebih'] for tahun in labels]
    buruk_data = [data_by_year[tahun]['buruk'] for tahun in labels]

    chart_data = {
        'labels': labels,
        'datasets': [
            {
                'label': 'Normal',
                'data': normal_data,
                'backgroundColor': 'rgba(75, 192, 192, 0.6)',
                'borderColor': 'rgba(75, 192, 192, 1)',
                'borderWidth': 1
            },
            {
                'label': 'Kurang',
                'data': kurang_data,
                'backgroundColor': 'rgba(255, 99, 132, 0.6)',
                'borderColor': 'rgba(255, 99, 132, 1)',
                'borderWidth': 1
            },
            {
                'label': 'Lebih',
                'data': lebih_data,
                'backgroundColor': 'rgba(255, 206, 86, 0.6)',
                'borderColor': 'rgba(255, 206, 86, 1)',
                'borderWidth': 1
            },
            {
                'label': 'Buruk',
                'data': buruk_data,
                'backgroundColor': 'rgba(54, 162, 235, 0.6)',
                'borderColor': 'rgba(54, 162, 235, 1)',
                'borderWidth': 1
            }
        ]
    }

    conn.close()
    return render_template('home.html', chart_data=chart_data)


# =============================================
# ROUTE AUTENTIKASI
# =============================================

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()
        if not username or not password:
            flash('Nama pengguna dan kata sandi harus diisi.', 'danger')
            return redirect(url_for('login'))
        conn = None
        try:
            conn = get_db()
            user = conn.execute('SELECT * FROM users WHERE username = ?', (username,)).fetchone()
            if user and check_password_hash(user['password'], password):
                session['user_id'] = user['id']
                session['username'] = user['username']
                session['role'] = user['role']
                session['name'] = user['nama_lengkap']  # Tambahkan ini
                session['photo'] = user['photo'] or 'default-profile.png'  # Tambahkan ini
                session['email'] = user['email'] if 'email' in user.keys() else ''
                flash('Login berhasil.', 'success')
                return redirect(url_for('dashboard'))
            else:
                flash('Nama pengguna atau kata sandi salah.', 'danger')
        except Exception as e:
            flash(f"Terjadi kesalahan: {str(e)}", 'danger')
        finally:
            if conn:
                conn.close()
    return render_template('auth/auth.html', page_mode='login')

@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()
        nama_lengkap = request.form.get('nama_lengkap', '').strip()
        
        if not username or not password or not nama_lengkap:
            flash('Semua field harus diisi', 'danger')
            return redirect(url_for('register'))
        
        conn = None  # Pastikan conn diinisialisasi
        try:
            conn = get_db()
            conn.execute('''
                INSERT INTO users (username, password, nama_lengkap, role)
                VALUES (?, ?, ?, ?)
            ''', (username, generate_password_hash(password), nama_lengkap, 'user'))
            conn.commit()
            flash('Registrasi berhasil! Silakan login', 'success')
            return redirect(url_for('login'))
        except sqlite3.IntegrityError:
            flash('Username sudah digunakan', 'danger')
        except Exception as e:
            flash(f"Terjadi kesalahan: {str(e)}", 'danger')
        finally:
            if conn:
                conn.close()
    return render_template('auth/auth.html', page_mode='register')

@app.route('/logout')
def logout():
    session.clear()
    flash('Anda telah logout', 'info')
    return redirect(url_for('login'))

# =============================================
# ROUTE DASHBOARD
# =============================================
@app.route('/dashboard', methods=['GET'])
@login_required
def dashboard():
    with get_db() as conn:
        stats = {
            'balita_count': conn.execute("SELECT COUNT(*) FROM balita").fetchone()[0],
            'status_gizi': conn.execute('''
                SELECT status_gizi, COUNT(*) as jumlah 
                FROM klasifikasi 
                GROUP BY status_gizi
            ''').fetchall()
        }

        # Data tren status gizi selama 6 bulan terakhir
        trend_data = conn.execute('''
            SELECT strftime('%Y-%m', tanggal_ukur) as bulan, 
                   status_gizi, COUNT(*) as jumlah
            FROM pengukuran p
            JOIN klasifikasi k ON p.id = k.pengukuran_id
            WHERE p.tanggal_ukur >= date('now', '-6 months')
            GROUP BY bulan, status_gizi
            ORDER BY bulan
        ''').fetchall()

        # Format data untuk Chart.js
        stats_gizi = {'normal':0, 'kurang':0, 'buruk':0}
        for row in stats['status_gizi']:
            stats_gizi[row['status_gizi']] = row['jumlah']

        # Data pengukuran terbaru (misal 5 terakhir)
        pengukuran_terbaru = conn.execute('''
            SELECT b.nama, b.nik, p.tanggal_ukur, p.berat_badan, p.tinggi_badan, k.status_gizi
            FROM pengukuran p
            JOIN balita b ON p.balita_id = b.id
            JOIN klasifikasi k ON p.id = k.pengukuran_id
            ORDER BY p.tanggal_ukur DESC
            LIMIT 10
        ''').fetchall()

    return render_template(
        'dashboard.html',
        balita_count=stats['balita_count'],
        stats_gizi=stats_gizi,
        pengukuran_terbaru=pengukuran_terbaru
    )
    
    
# =============================================
# ROUTE KELOLA PROFILE
# =============================================
@app.route('/profile')
@login_required
def profile():
    conn = get_db()
    user = conn.execute('SELECT * FROM users WHERE id = ?', (session['user_id'],)).fetchone()
    conn.close()
    return render_template('profile.html', user=user)

@app.route('/profile/edit', methods=['GET', 'POST'])
@login_required
def edit_profil():
    conn = get_db()
    user = conn.execute('SELECT * FROM users WHERE id = ?', (session['user_id'],)).fetchone()
    if request.method == 'POST':
        nama_lengkap = request.form['nama_lengkap']
        alamat = request.form['alamat']
        tempat_lahir = request.form['tempat_lahir']
        tanggal_lahir = request.form['tanggal_lahir']
        pendidikan = request.form['pendidikan']
        email = request.form['email']

        # Foto profil
        photo = user['photo']
        if 'photo' in request.files:
            file = request.files['photo']
            if file and file.filename:
                filename = f"user_{session['user_id']}.png"
                filepath = os.path.join(app.root_path, 'static/img', filename)
                file.save(filepath)
                photo = filename

        conn.execute('''
            UPDATE users SET 
                nama_lengkap=?, alamat=?, tempat_lahir=?, tanggal_lahir=?, pendidikan=?, email=?, photo=?
            WHERE id=?
        ''', (nama_lengkap, alamat, tempat_lahir, tanggal_lahir, pendidikan, email, photo, session['user_id']))
        conn.commit()
        conn.close()
        # Update session
        session['name'] = nama_lengkap
        session['photo'] = photo
        session['email'] = email
        flash('Profil berhasil diperbarui', 'success')
        return redirect(url_for('profile'))
    conn.close()
    return render_template('edit_profil.html', user=user)
    
# =============================================
# ROUTE KELOLA USER
# =============================================

@app.route('/user')
@admin_required
def kelola_user():
    conn = get_db()
    users = conn.execute('SELECT * FROM users').fetchall()
    conn.close()
    return render_template('user/index.html', users=users)

@app.route('/user/tambah', methods=['GET', 'POST'])
@admin_required
def tambah_user():
    if request.method == 'POST':
        username = request.form['username']
        nama_lengkap = request.form['nama_lengkap']
        password = generate_password_hash(request.form['password'])
        role = request.form['role']
        conn = get_db()
        try:
            conn.execute('INSERT INTO users (username, nama_lengkap, password, role) VALUES (?, ?, ?, ?)', 
                         (username, nama_lengkap, password, role))
            conn.commit()
            flash('User berhasil ditambah', 'success')
        except sqlite3.IntegrityError:
            flash('Username sudah digunakan', 'danger')
        conn.close()
        return redirect(url_for('kelola_user'))
    return render_template('user/tambah.html')

@app.route('/user/edit/<int:id>', methods=['GET', 'POST'])
@admin_required
def edit_user(id):
    conn = get_db()
    user = conn.execute('SELECT * FROM users WHERE id = ?', (id,)).fetchone()
    if request.method == 'POST':
        username = request.form['username']
        nama_lengkap = request.form['nama_lengkap']
        role = request.form['role']
        if request.form['password']:
            password = generate_password_hash(request.form['password'])
            conn.execute('UPDATE users SET username=?, nama_lengkap=?, password=?, role=? WHERE id=?',
                         (username, nama_lengkap, password, role, id))
        else:
            conn.execute('UPDATE users SET username=?, nama_lengkap=?, role=? WHERE id=?',
                         (username, nama_lengkap, role, id))
        conn.commit()
        conn.close()
        flash('User berhasil diupdate', 'success')
        return redirect(url_for('kelola_user'))
    conn.close()
    return render_template('user/edit.html', user=user)

@app.route('/user/hapus/<int:id>', methods=['POST'])
@admin_required
def hapus_user(id):
    conn = get_db()
    conn.execute('DELETE FROM users WHERE id = ?', (id,))
    conn.commit()
    conn.close()
    flash('User berhasil dihapus', 'success')
    return redirect(url_for('kelola_user'))

# =============================================
# ROUTE KELOLA BALITA
# =============================================

@app.route('/balita')
@login_required
def data_balita():
    """Endpoint untuk melihat data balita dengan filter dan pagination"""
    # Ambil parameter filter dari query string
    nama = request.args.get('nama', '').strip()
    tanggal_mulai = request.args.get('tanggal_mulai', '').strip()
    tanggal_akhir = request.args.get('tanggal_akhir', '').strip()
    page = request.args.get('page', 1, type=int)
    per_page = 50
    offset = (page - 1) * per_page

    conn = get_db()
    try:
        # Query utama
        query = '''
            SELECT b.*, 
                   (SELECT COUNT(*) FROM pengukuran WHERE balita_id = b.id) as jumlah_pengukuran,
                   CAST((julianday('now') - julianday(b.tanggal_lahir)) / 365 AS INTEGER) as usia_tahun,
                   CAST(((julianday('now') - julianday(b.tanggal_lahir)) % 365) / 30 AS INTEGER) as usia_bulan
            FROM balita b
            WHERE 1=1
        '''
        params = []

        if nama:
            query += ' AND b.nama LIKE ?'
            params.append(f'%{nama}%')
        if tanggal_mulai:
            query += ' AND b.tanggal_lahir >= ?'
            params.append(tanggal_mulai)
        if tanggal_akhir:
            query += ' AND b.tanggal_lahir <= ?'
            params.append(tanggal_akhir)

        query += ' ORDER BY b.nama LIMIT ? OFFSET ?'
        params.extend([per_page, offset])

        # Eksekusi query utama
        balita = conn.execute(query, params).fetchall()

        # Query untuk menghitung total data
        count_query = '''
            SELECT COUNT(*)
            FROM balita b
            WHERE 1=1
        '''
        count_params = []

        if nama:
            count_query += ' AND b.nama LIKE ?'
            count_params.append(f'%{nama}%')
        if tanggal_mulai:
            count_query += ' AND b.tanggal_lahir >= ?'
            count_params.append(tanggal_mulai)
        if tanggal_akhir:
            count_query += ' AND b.tanggal_lahir <= ?'
            count_params.append(tanggal_akhir)

        # Eksekusi query hitung total data
        if count_params:
            total_data = conn.execute(count_query, count_params).fetchone()[0]
        else:
            total_data = conn.execute(count_query).fetchone()[0]

    finally:
        conn.close()

    # Tentukan apakah ada halaman berikutnya
    has_next = (offset + per_page) < total_data

    return render_template('balita/data.html', balita=balita, page=page, has_next=has_next)

def hitung_usia(tanggal_lahir):
    lahir = datetime.strptime(tanggal_lahir, '%Y-%m-%d')
    sekarang = datetime.now()

    tahun = sekarang.year - lahir.year
    bulan = sekarang.month - lahir.month

    if bulan < 0:
        tahun -= 1
        bulan += 12

    return tahun, bulan

@app.route('/balita/tambah', methods=['GET', 'POST'])
def tambah_balita():
    if request.method == 'POST':
        nama = request.form['nama'].strip()
        nik = request.form.get('nik', '').strip()
        tanggal_lahir = request.form['tanggal_lahir'].strip()
        jenis_kelamin = request.form['jenis_kelamin']
        nama_ortu = request.form['nama_ortu'].strip()

        # Validasi NIK
        if not re.match(r'^\d{16}$', nik):
            flash('NIK harus 16 digit angka', 'danger')
            return redirect(url_for('tambah_balita'))

        conn = get_db()
        try:
            # Cek duplikasi NIK (opsional)
            existing = conn.execute('SELECT id FROM balita WHERE nik = ?', (nik,)).fetchone()
            if existing:
                flash('NIK sudah terdaftar!', 'danger')
                return redirect(url_for('tambah_balita'))

            conn.execute(
                'INSERT INTO balita (nama, nik, tanggal_lahir, jenis_kelamin, nama_ortu) VALUES (?, ?, ?, ?, ?)',
                (nama, nik, tanggal_lahir, jenis_kelamin, nama_ortu)
            )
            conn.commit()
            flash('Balita berhasil ditambahkan', 'success')
            return redirect(url_for('data_balita'))
        except Exception as e:
            conn.rollback()
            flash(f'Gagal menambah balita: {e}', 'danger')
            return redirect(url_for('tambah_balita'))
        finally:
            conn.close()
    return render_template('balita/tambah.html')

@app.route('/balita/edit/<int:id>', methods=['GET', 'POST'])
def edit_balita(id):
    conn = get_db()
    balita = conn.execute('SELECT * FROM balita WHERE id = ?', (id,)).fetchone()
    if not balita:
        flash('Data balita tidak ditemukan', 'danger')
        return redirect(url_for('data_balita'))
    if request.method == 'POST':
        nama = request.form['nama'].strip()
        nik = request.form.get('nik', '').strip()
        tanggal_lahir = request.form['tanggal_lahir'].strip()
        jenis_kelamin = request.form['jenis_kelamin']
        nama_ortu = request.form['nama_ortu'].strip()

        if not re.match(r'^\d{16}$', nik):
            flash('NIK harus 16 digit angka', 'danger')
            return redirect(url_for('edit_balita', id=id))

        # Cek duplikasi NIK selain milik balita ini
        existing = conn.execute('SELECT id FROM balita WHERE nik = ? AND id != ?', (nik, id)).fetchone()
        if existing:
            flash('NIK sudah terdaftar di balita lain!', 'danger')
            return redirect(url_for('edit_balita', id=id))

        try:
            conn.execute(
                'UPDATE balita SET nama=?, nik=?, tanggal_lahir=?, jenis_kelamin=?, nama_ortu=? WHERE id=?',
                (nama, nik, tanggal_lahir, jenis_kelamin, nama_ortu, id)
            )
            conn.commit()
            flash('Data balita berhasil diupdate', 'success')
            return redirect(url_for('data_balita'))
        except Exception as e:
            conn.rollback()
            flash(f'Gagal update balita: {e}', 'danger')
            return redirect(url_for('edit_balita', id=id))
        finally:
            conn.close()
    conn.close()
    return render_template('balita/edit.html', balita=balita)

@app.route('/balita/hapus/<int:id>')
@login_required
def hapus_balita(id):
    conn = get_db()
    try:
        conn.execute('DELETE FROM balita WHERE id = ?', (id,))
        conn.commit()
        flash('Data balita berhasil dihapus', 'success')
    except Exception as e:
        conn.rollback()
        flash(f'Error: {str(e)}', 'danger')
    finally:
        conn.close()
    
    return redirect(url_for('data_balita'))


# =============================================
# ROUTE KELOLA PENGUKURAN
# =============================================

@app.route('/pengukuran')
@login_required
def data_pengukuran():
    """Endpoint untuk melihat data pengukuran dengan filter dan pagination"""
    # Ambil parameter filter dari query string
    nama_balita = request.args.get('nama_balita', '').strip()
    status_gizi = request.args.get('status_gizi', '').strip()
    page = request.args.get('page', 1, type=int)
    per_page = 50
    offset = (page - 1) * per_page

    conn = get_db()
    try:
        # Query utama
        query = '''
            SELECT p.*, b.nama as nama_balita, b.nik as nik_balita, k.status_gizi
            FROM pengukuran p
            JOIN balita b ON p.balita_id = b.id
            LEFT JOIN klasifikasi k ON p.id = k.pengukuran_id
            WHERE 1=1
        '''
        params = []

        if nama_balita:
            query += ' AND b.nama LIKE ?'
            params.append(f'%{nama_balita}%')
        if status_gizi:
            query += ' AND k.status_gizi = ?'
            params.append(status_gizi)

        # Tambahkan pagination
        query += ' ORDER BY p.tanggal_ukur DESC LIMIT ? OFFSET ?'
        params.extend([per_page, offset])

        # Eksekusi query utama
        pengukuran = conn.execute(query, params).fetchall()

        # Query untuk menghitung total data
        count_query = '''
            SELECT COUNT(*)
            FROM pengukuran p
            JOIN balita b ON p.balita_id = b.id
            LEFT JOIN klasifikasi k ON p.id = k.pengukuran_id
            WHERE 1=1
        '''
        count_params = []

        if nama_balita:
            count_query += ' AND b.nama LIKE ?'
            count_params.append(f'%{nama_balita}%')
        if status_gizi:
            count_query += ' AND k.status_gizi = ?'
            count_params.append(status_gizi)

        # Eksekusi query hitung total data
        if count_params:
            total_data = conn.execute(count_query, count_params).fetchone()[0]
        else:
            total_data = conn.execute(count_query).fetchone()[0]

    finally:
        conn.close()

    # Tentukan apakah ada halaman berikutnya
    has_next = (offset + per_page) < total_data

    return render_template('pengukuran/data.html', pengukuran=pengukuran, page=page, has_next=has_next)

def klasifikasi_knn(berat_badan, tinggi_badan, lingkar_lengan):
    conn = get_db()
    # ambil dari dataset_lvq!
    data_protos = conn.execute('SELECT feature1, feature2, feature3, target FROM dataset_lvq').fetchall()
    if not data_protos:
        return "Dataset LVQ kosong, lakukan sampling LVQ dahulu"
    X_train = [[row['feature1'], row['feature2'], row['feature3']] for row in data_protos]
    y_train = [row['target'] for row in data_protos]
    params = conn.execute('SELECT * FROM parameter_knn').fetchone()
    nilai_k = params['nilai_k'] if params else 3
    bobot = (
        float(params['bb']) if 'bb' in params.keys() else 1,
        float(params['tb']) if 'tb' in params.keys() else 1,
        float(params['lila']) if 'lila' in params.keys() else 1
    )
    model = KNN(k=nilai_k, bobot=bobot)
    model.fit(X_train, y_train)
    fitur_balita = [[berat_badan, tinggi_badan, lingkar_lengan]]
    status_gizi = model.predict(fitur_balita)[0]
    return status_gizi

        
@app.route('/pengukuran/tambah', methods=['GET', 'POST'])
@login_required
def tambah_pengukuran():
    if request.method == 'POST':
        balita_id = request.form.get('balita_id', '')
        berat_badan = request.form.get('berat_badan', '')
        tinggi_badan = request.form.get('tinggi_badan', '')
        lingkar_lengan = request.form.get('lingkar_lengan', '')

        # Validasi input
        if not all([balita_id, berat_badan, tinggi_badan, lingkar_lengan]):
            flash('Semua field harus diisi', 'danger')
            return redirect(url_for('tambah_pengukuran'))

        try:
            # Konversi input ke tipe data yang sesuai
            berat_badan = float(berat_badan)
            tinggi_badan = float(tinggi_badan)
            lingkar_lengan = float(lingkar_lengan)

            # Operasi database
            with get_db() as conn:
                balita = conn.execute(
                    'SELECT tanggal_lahir, jenis_kelamin FROM balita WHERE id = ?', 
                    (balita_id,)
                ).fetchone()

                if not balita:
                    flash('Data balita tidak ditemukan', 'danger')
                    return redirect(url_for('tambah_pengukuran'))

                # Prediksi status gizi menggunakan KNN
                status_gizi = klasifikasi_knn(berat_badan, tinggi_badan, lingkar_lengan)
                if status_gizi == "Data Training Kosong":
                    flash('Data training belum tersedia, tidak bisa melakukan klasifikasi!', 'danger')
                    return redirect(url_for('tambah_pengukuran'))

                # Simpan data pengukuran
                conn.execute('''
                    INSERT INTO pengukuran 
                    (balita_id, tanggal_ukur, berat_badan, tinggi_badan, lingkar_lengan)
                    VALUES (?, date('now'), ?, ?, ?)
                ''', (balita_id, berat_badan, tinggi_badan, lingkar_lengan))

                pengukuran_id = conn.execute('SELECT last_insert_rowid()').fetchone()[0]

                # Simpan hasil klasifikasi
                conn.execute('''
                    INSERT INTO klasifikasi 
                    (pengukuran_id, status_gizi, tanggal_klasifikasi)
                    VALUES (?, ?, date('now'))
                ''', (pengukuran_id, status_gizi))

                conn.commit()
                flash(f'Pengukuran berhasil disimpan. Status gizi: {status_gizi}', 'success')
                return redirect(url_for('data_pengukuran'))

        except ValueError:
            flash('Input tidak valid. Pastikan angka yang dimasukkan benar', 'danger')
        except Exception as e:
            flash(f'Error: {str(e)}', 'danger')

    # Render form tambah pengukuran
    with get_db() as conn:
        balita = conn.execute('SELECT id, nama FROM balita ORDER BY nama').fetchall()
    return render_template('pengukuran/tambah.html', balita=balita)

@app.route('/pengukuran/edit/<int:id>', methods=['GET', 'POST'])
@login_required
def edit_pengukuran(id):
    conn = get_db()
    try:
        if request.method == 'POST':
            berat_badan = request.form.get('berat_badan', '')
            tinggi_badan = request.form.get('tinggi_badan', '')
            lingkar_lengan = request.form.get('lingkar_lengan', '')
            if not all([berat_badan, tinggi_badan, lingkar_lengan]):
                flash('Semua field harus diisi', 'danger')
                return redirect(url_for('edit_pengukuran', id=id))
            try:
                berat_badan = float(berat_badan)
                tinggi_badan = float(tinggi_badan)
                lingkar_lengan = float(lingkar_lengan)
                status_gizi = klasifikasi_knn(berat_badan, tinggi_badan, lingkar_lengan)
                if status_gizi == "Data Training Kosong":
                    flash('Data training belum tersedia, tidak bisa klasifikasi!', 'danger')
                    return redirect(url_for('edit_pengukuran', id=id))
                conn.execute('''
                    UPDATE pengukuran 
                    SET berat_badan = ?, tinggi_badan = ?, lingkar_lengan = ?
                    WHERE id = ?
                ''', (berat_badan, tinggi_badan, lingkar_lengan, id))
                conn.execute('''
                    UPDATE klasifikasi
                    SET status_gizi = ?, tanggal_klasifikasi = date('now')
                    WHERE pengukuran_id = ?
                ''', (status_gizi, id))
                conn.commit()
                flash('Data pengukuran & status gizi berhasil diupdate', 'success')
                return redirect(url_for('data_pengukuran'))
            except ValueError:
                flash('Input tidak valid. Pastikan angka yang dimasukkan benar', 'danger')
            except Exception as e:
                conn.rollback()
                flash(f'Error: {str(e)}', 'danger')
        pengukuran = conn.execute('''
            SELECT p.*, b.nama, b.tanggal_lahir, b.jenis_kelamin
            FROM pengukuran p
            JOIN balita b ON p.balita_id = b.id
            WHERE p.id = ?
        ''', (id,)).fetchone()
        if not pengukuran:
            flash('Data pengukuran tidak ditemukan', 'danger')
            return redirect(url_for('data_pengukuran'))
        return render_template('pengukuran/edit.html', pengukuran=pengukuran)
    finally:
        conn.close()

@app.route('/pengukuran/hapus/<int:id>')
@login_required
def hapus_pengukuran(id):
    conn = get_db()
    try:
        conn.execute('DELETE FROM klasifikasi WHERE pengukuran_id = ?', (id,))
        conn.execute('DELETE FROM pengukuran WHERE id = ?', (id,))
        conn.commit()
        flash('Data pengukuran berhasil dihapus', 'success')
    except Exception as e:
        conn.rollback()
        flash(f'Error: {str(e)}', 'danger')
    finally:
        conn.close()
    return redirect(url_for('data_pengukuran'))

# =============================================
# ROUTER KELOAL LAPORAN
# =============================================

@app.route('/laporan')
@login_required
def laporan():
    conn = get_db()
    
    # Query untuk data grafik trend (semua data)
    trend_query = '''
        SELECT strftime('%Y-%m', tanggal_ukur) as bulan,
               k.status_gizi,
               COUNT(*) as jumlah
        FROM pengukuran p
        JOIN klasifikasi k ON p.id = k.pengukuran_id
        WHERE k.status_gizi IN ('normal', 'kurang', 'buruk')
        GROUP BY bulan, k.status_gizi
        ORDER BY bulan ASC
    '''
    
    trend_data = conn.execute(trend_query).fetchall()
    
    # Query untuk data dalam tabel
    data_query = '''
        SELECT b.nama, b.nik, p.tanggal_ukur, p.berat_badan, p.tinggi_badan, 
               p.lingkar_lengan, k.status_gizi
        FROM pengukuran p
        JOIN balita b ON p.balita_id = b.id
        JOIN klasifikasi k ON p.id = k.pengukuran_id
        ORDER BY p.tanggal_ukur ASC
    '''
    
    data = conn.execute(data_query).fetchall()
    
    # Proses data untuk grafik
    months = sorted(set(row['bulan'] for row in trend_data))
    status_types = ['normal', 'kurang', 'buruk']  # Hanya tampilkan 3 status ini
    
    # Warna untuk setiap status
    colors = {
        'normal': '#28a745',
        'kurang': '#ffc107',
        'buruk': '#dc3545'
    }
    
    # Dataset untuk setiap status
    datasets = []
    for status in status_types:
        data_status = []
        for month in months:
            count = next(
                (row['jumlah'] for row in trend_data 
                 if row['bulan'] == month and row['status_gizi'] == status),
                0
            )
            data_status.append(count)
        
        datasets.append({
            'label': status.title(),
            'data': data_status,
            'color': colors.get(status.lower(), '#000000')
        })
    
    conn.close()
    
    return render_template(
        'laporan/laporan.html',
        data=data,
        labels=months,
        datasets=datasets
    )

@app.route('/laporan/data')
@login_required
def laporan_data():
    try:
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        
        if not start_date or not end_date:
            return jsonify({'error': 'Tanggal awal dan akhir harus diisi'}), 400
            
        conn = get_db()
        
        # Query untuk data tabel
        query_table = '''
            SELECT b.nama, b.nik, p.tanggal_ukur, p.berat_badan, p.tinggi_badan, 
                   p.lingkar_lengan, k.status_gizi
            FROM pengukuran p
            JOIN balita b ON p.balita_id = b.id
            JOIN klasifikasi k ON p.id = k.pengukuran_id
            WHERE p.tanggal_ukur BETWEEN ? AND ?
            ORDER BY p.tanggal_ukur ASC
        '''
        
        # Query untuk data trend sesuai filter
        query_trend = '''
            SELECT strftime('%Y-%m', tanggal_ukur) as bulan,
                   k.status_gizi,
                   COUNT(*) as jumlah
            FROM pengukuran p
            JOIN klasifikasi k ON p.id = k.pengukuran_id
            WHERE p.tanggal_ukur BETWEEN ? AND ?
            AND k.status_gizi IN ('normal', 'kurang', 'buruk')
            GROUP BY bulan, k.status_gizi
            ORDER BY bulan ASC
        '''
        
        table_data = conn.execute(query_table, [start_date, end_date]).fetchall()
        trend_data = conn.execute(query_trend, [start_date, end_date]).fetchall()
        
        # Proses data tabel
        result_table = []
        for row in table_data:
            result_table.append({
                'nik': row['nik'],
                'nama': row['nama'],
                'tanggal_ukur': row['tanggal_ukur'],
                'berat_badan': row['berat_badan'],
                'tinggi_badan': row['tinggi_badan'],
                'lingkar_lengan': row['lingkar_lengan'],
                'status_gizi': row['status_gizi']
            })
        
        # Proses data trend
        months = sorted(set(row['bulan'] for row in trend_data))
        status_types = ['normal', 'kurang', 'buruk']
        
        colors = {
            'normal': '#28a745',
            'kurang': '#ffc107',
            'buruk': '#dc3545'
        }
        
        datasets = []
        for status in status_types:
            data_status = []
            for month in months:
                count = next(
                    (row['jumlah'] for row in trend_data 
                     if row['bulan'] == month and row['status_gizi'] == status),
                    0
                )
                data_status.append(count)
            
            datasets.append({
                'label': status.title(),
                'data': data_status,
                'borderColor': colors.get(status.lower(), '#000000'),
                'backgroundColor': colors.get(status.lower(), '#000000')
            })
        
        conn.close()
        
        return jsonify({
            'table_data': result_table,
            'chart_data': {
                'labels': months,
                'datasets': datasets
            }
        })
        
    except Exception as e:
        print(f"Error in laporan_data: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/laporan/pdf')
@login_required
def laporan_pdf():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    query = '''
        SELECT b.nama, b.nik, p.tanggal_ukur, p.berat_badan, p.tinggi_badan, p.lingkar_lengan, k.status_gizi
        FROM pengukuran p
        JOIN balita b ON p.balita_id = b.id
        JOIN klasifikasi k ON p.id = k.pengukuran_id
    '''
    
    params = []
    if start_date and end_date:
        query += ' WHERE p.tanggal_ukur BETWEEN ? AND ?'
        params.extend([start_date, end_date])
    elif start_date:
        query += ' WHERE p.tanggal_ukur >= ?'
        params.append(start_date)
    elif end_date:
        query += ' WHERE p.tanggal_ukur <= ?'
        params.append(end_date)
        
    query += ' ORDER BY p.tanggal_ukur DESC'
    
    conn = get_db()
    if params:
        data = conn.execute(query, params).fetchall()
    else:
        data = conn.execute(query).fetchall()
    conn.close()
    
    rendered = render_template('laporan/pdf_template.html', 
                             data=data,
                             start_date=start_date,
                             end_date=end_date)
    pdf = pdfkit.from_string(rendered, False)
    response = make_response(pdf)
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = 'attachment; filename=laporan.pdf'
    return response

@app.route('/laporan/excel')
@login_required
def laporan_excel():
    try:
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        
        query = '''
            SELECT b.nama, b.nik, p.tanggal_ukur, p.berat_badan, p.tinggi_badan, 
                   p.lingkar_lengan, k.status_gizi
            FROM pengukuran p
            JOIN balita b ON p.balita_id = b.id
            JOIN klasifikasi k ON p.id = k.pengukuran_id
        '''
        params = []
        
        if start_date and end_date:
            query += ' WHERE p.tanggal_ukur BETWEEN ? AND ?'
            params.extend([start_date, end_date])
        elif start_date:
            query += ' WHERE p.tanggal_ukur >= ?'
            params.append(start_date)
        elif end_date:
            query += ' WHERE p.tanggal_ukur <= ?'
            params.append(end_date)
            
        query += ' ORDER BY p.tanggal_ukur ASC'
        
        conn = get_db()
        if params:
            data = conn.execute(query, params).fetchall()
        else:
            data = conn.execute(query).fetchall()
        conn.close()

        # Buat workbook baru
        wb = Workbook()
        ws = wb.active
        ws.title = "Laporan Status Gizi"

        # Styling
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        
        # Judul laporan
        periode = f"Data Pengukuran"
        if start_date and end_date:
            periode += f" ({start_date} - {end_date})"
        ws['A1'] = periode
        ws.merge_cells('A1:H1')
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')

        # Header tabel
        headers = ['No', 'NIK', 'Nama Balita', 'Tanggal Ukur', 'Berat Badan (kg)', 
                  'Tinggi Badan (cm)', 'Lingkar Lengan (cm)', 'Status Gizi']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill

        # Isi data
        for row_idx, item in enumerate(data, 1):
            ws.cell(row=row_idx+3, column=1).value = row_idx
            ws.cell(row=row_idx+3, column=2).value = item['nik']
            ws.cell(row=row_idx+3, column=3).value = item['nama']
            ws.cell(row=row_idx+3, column=4).value = item['tanggal_ukur']
            ws.cell(row=row_idx+3, column=5).value = item['berat_badan']
            ws.cell(row=row_idx+3, column=6).value = item['tinggi_badan']
            ws.cell(row=row_idx+3, column=7).value = item['lingkar_lengan']
            ws.cell(row=row_idx+3, column=8).value = item['status_gizi']

        # Auto-adjust column width
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        # Simpan ke BytesIO
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='laporan_gizi_balita.xlsx'
        )

    except Exception as e:
        return str(e), 500

@app.route('/laporan/cetak')
@login_required
def cetak_laporan():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    query = '''
        SELECT b.nama, b.nik, p.tanggal_ukur, p.berat_badan, p.tinggi_badan, p.lingkar_lengan, k.status_gizi
        FROM pengukuran p
        JOIN balita b ON p.balita_id = b.id
        JOIN klasifikasi k ON p.id = k.pengukuran_id
    '''
    
    params = []
    if start_date and end_date:
        query += ' WHERE p.tanggal_ukur BETWEEN ? AND ?'
        params.extend([start_date, end_date])
    elif start_date:
        query += ' WHERE p.tanggal_ukur >= ?'
        params.append(start_date)
    elif end_date:
        query += ' WHERE p.tanggal_ukur <= ?'
        params.append(end_date)
        
    query += ' ORDER BY p.tanggal_ukur DESC'
    
    conn = get_db()
    if params:
        data = conn.execute(query, params).fetchall()
    else:
        data = conn.execute(query).fetchall()
    conn.close()
    
    return render_template('laporan/cetak.html', 
                         data=data,
                         start_date=start_date,
                         end_date=end_date)

# =============================================
# ROUTE PARAMETER KNN (ADMIN ONLY)
# =============================================

def evaluasi_model_with_parameters(nilai_k, bobot):
    conn = get_db()
    try:
        # Ambil data hasil sampling LVQ
        data = conn.execute('''
            SELECT feature1, feature2, feature3, target
            FROM dataset_lvq
        ''').fetchall()
        
        if not data:
            raise Exception("Dataset hasil sampling LVQ kosong!")

        # Siapkan data
        X = np.array([[row['feature1'], row['feature2'], row['feature3']] for row in data])
        y = np.array([row['target'] for row in data])

        # Pastikan y mengandung semua kelas yang diharapkan
        class_names = ['normal', 'kurang', 'buruk']
        if not all(cls in np.unique(y) for cls in class_names):
            raise Exception("Dataset tidak memiliki semua kelas yang diperlukan (normal, kurang, buruk)")

        # Inisialisasi model dengan 3 bobot pertama
        model = KNN(k=nilai_k, bobot=bobot[:3])

        # Cross validation 
        cv = StratifiedKFold(n_splits=5, shuffle=True, random_state=42)
        cv_scores = cross_val_score(model, X, y, cv=cv, scoring='accuracy')
        cv_scores = cv_scores * 100

        # Split data untuk evaluasi final
        skf = StratifiedKFold(n_splits=5, shuffle=True, random_state=42)
        train_idx, test_idx = next(skf.split(X, y))
        X_train, X_test = X[train_idx], X[test_idx]
        y_train, y_test = y[train_idx], y[test_idx]

        # Fit dan prediksi
        model.fit(X_train, y_train)
        y_pred = model.predict(X_test)
        
        # Hitung metrik
        accuracy = accuracy_score(y_test, y_pred) * 100
        matrix = confusion_matrix(y_test, y_pred, labels=class_names)
        
        # Hitung metrik per kelas
        n_classes = len(class_names)
        sensitivitas = []
        spesifisitas = []
        presisi = []
        f1_scores = []
        
        for i in range(n_classes):
            # True Positive, False Positive, False Negative, True Negative
            tp = matrix[i, i]
            fp = np.sum(matrix[:, i]) - tp
            fn = np.sum(matrix[i, :]) - tp
            tn = np.sum(matrix) - (tp + fp + fn)
            
            # Hitung metrik
            sens = tp / (tp + fn) if (tp + fn) > 0 else 0
            spec = tn / (tn + fp) if (tn + fp) > 0 else 0
            prec = tp / (tp + fp) if (tp + fp) > 0 else 0
            f1 = 2 * (prec * sens) / (prec + sens) if (prec + sens) > 0 else 0
            
            sensitivitas.append(sens)
            spesifisitas.append(spec)
            presisi.append(prec)
            f1_scores.append(f1)
        
        return {
            'accuracy': accuracy,
            'mean_sensitivitas': np.mean(sensitivitas),
            'mean_spesifisitas': np.mean(spesifisitas),
            'cv_mean': np.mean(cv_scores),
            'cv_std': np.std(cv_scores),
            'confusion_matrix': matrix.tolist(),
            'sensitivitas_per_kelas': sensitivitas,
            'spesifisitas_per_kelas': spesifisitas,
            'presisi_per_kelas': presisi,
            'f1_per_kelas': f1_scores,
            'class_names': class_names  # Tambahkan nama kelas
        }
    except Exception as e:
        raise Exception(f"Error dalam evaluasi model: {str(e)}")
    finally:
        conn.close()

@app.route('/parameter', methods=['GET', 'POST'])
@login_required
def parameter():
    if request.method == 'POST':
        nilai_k = int(request.form['nilai_k'])
        bobot_berat = float(request.form['bobot_bb'])  # pastikan nama field form sesuai
        bobot_tinggi = float(request.form['bobot_tb'])
        bobot_lila = float(request.form['bobot_ll'])
        bobot_umur = float(request.form['bobot_umur'])
        bobot_jk = float(request.form['bobot_jk'])
        
        # Validasi total bobot
        total_bobot = bobot_berat + bobot_tinggi + bobot_lila + bobot_umur + bobot_jk
        if abs(total_bobot - 1.0) > 0.0001:
            flash('Total bobot harus sama dengan 1.0', 'danger')
            return redirect(url_for('parameter'))
            
        conn = get_db()
        try:
            cur = conn.execute(
                '''INSERT INTO parameter_knn 
                   (nilai_k, bobot_berat, bobot_tinggi, bobot_lila, bobot_umur, bobot_jk, created_at) 
                   VALUES (?, ?, ?, ?, ?, ?, datetime('now'))''',
                (nilai_k, bobot_berat, bobot_tinggi, bobot_lila, bobot_umur, bobot_jk)
            )
            parameter_id = cur.lastrowid

            conn.execute(
                '''INSERT INTO parameter_history 
                   (parameter_id, changed_by, nilai_k, bobot_berat, bobot_tinggi, bobot_lila, 
                    bobot_umur, bobot_jk, changed_at) 
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, datetime('now'))''',
                (parameter_id, session['user_id'], nilai_k, bobot_berat, bobot_tinggi, 
                 bobot_lila, bobot_umur, bobot_jk)
            )
            
            # Evaluasi model
            hasil = evaluasi_model_with_parameters(
                nilai_k,
                (bobot_berat, bobot_tinggi, bobot_lila, bobot_umur, bobot_jk)
            )
            
            session['last_evaluation'] = hasil
            conn.commit()
            flash(f'Parameter berhasil disimpan dengan akurasi: {hasil["accuracy"]:.2f}%', 'success')
            
        except Exception as e:
            conn.rollback()
            flash(f'Error: {str(e)}', 'error')
        finally:
            conn.close()
            
        return redirect(url_for('evaluasi_model'))

    # GET request
    conn = get_db()
    params = conn.execute('SELECT * FROM parameter_knn ORDER BY created_at DESC LIMIT 1').fetchone()
    history = None
    if params:
        history = conn.execute('''
            SELECT ph.*, u.username as pengubah
            FROM parameter_history ph
            JOIN users u ON ph.changed_by = u.id
            ORDER BY changed_at DESC LIMIT 10
        ''').fetchall()
    conn.close()
    
    return render_template('admin/parameter.html', params=params, history=history)

@app.route('/admin/parameter/update', methods=['POST'])
@admin_required
def update_parameter():
    # Validasi input
    try:
        nilai_k = int(request.form['nilai_k'])
        bobot_berat = float(request.form['bobot_berat'])
        bobot_tinggi = float(request.form['bobot_tinggi'])
        bobot_lila = float(request.form['bobot_lila'])
        bobot_umur = float(request.form['bobot_umur'])
        bobot_jk = float(request.form['bobot_jk'])
    except ValueError:
        flash('Input tidak valid', 'danger')
        return redirect(url_for('parameter'))
    
    # Validasi total bobot
    total_bobot = bobot_berat + bobot_tinggi + bobot_lila + bobot_umur + bobot_jk
    if not 0.99 <= total_bobot <= 1.01:
        flash('Total bobot harus sama dengan 1.0', 'danger')
        return redirect(url_for('parameter'))
    
    conn = get_db()
    try:
        # Update parameter
        conn.execute('''
            UPDATE parameter_knn 
            SET nilai_k = ?, 
                bobot_berat = ?, 
                bobot_tinggi = ?, 
                bobot_lila = ?, 
                bobot_umur = ?, 
                bobot_jk = ?
        ''', (nilai_k, bobot_berat, bobot_tinggi, bobot_lila, bobot_umur, bobot_jk))
        
        # Simpan history
        conn.execute('''
            INSERT INTO parameter_history 
            (parameter_id, changed_by, nilai_k, bobot_berat, bobot_tinggi, bobot_lila, bobot_umur, bobot_jk)
            VALUES (1, ?, ?, ?, ?, ?, ?, ?)
        ''', (session['user_id'], nilai_k, bobot_berat, bobot_tinggi, bobot_lila, bobot_umur, bobot_jk))
        
        conn.commit()
        flash('Parameter berhasil diperbarui', 'success')
    except sqlite3.OperationalError as e:
        conn.rollback()
        flash(f'Database error: {str(e)}', 'danger')
    except Exception as e:
        conn.rollback()
        flash(f'Error: {str(e)}', 'danger')
    finally:
        conn.close()
    
    return redirect(url_for('parameter'))

# =============================================
# ROUTE DATASET TRAINING (ADMIN ONLY)
# =============================================

@app.route('/kelola_dataset_training')
@login_required
def kelola_dataset_training():
    conn = get_db()
    data = conn.execute('SELECT id, feature1, feature2, feature3, target FROM dataset_training').fetchall()
    conn.close()
    return render_template('dataset/kelola_dataset_training.html', data=data)

@app.route('/hapus_dataset_training/<int:id>', methods=['POST'])
@login_required
def hapus_dataset_training(id):
    conn = get_db()
    conn.execute('DELETE FROM dataset_training WHERE id=?', (id,))
    conn.commit()
    conn.close()
    flash('Data training berhasil dihapus.', 'success')
    return redirect(url_for('kelola_dataset_training'))


@app.route('/tambah_dataset_training', methods=['GET', 'POST'])
@login_required
def tambah_dataset_training():
    if request.method == 'POST':
        bb = float(request.form['bb'])
        tb = float(request.form['tb'])
        lila = float(request.form['lila'])
        status = request.form['status']
        conn = get_db()
        conn.execute(
            'INSERT INTO dataset_training (feature1, feature2, feature3, target) VALUES (?, ?, ?, ?)',
            (bb, tb, lila, status)
        )
        conn.commit()
        conn.close()
        flash('Data training berhasil ditambahkan.', 'success')
        return redirect(url_for('kelola_dataset_training'))
    return render_template('dataset/tambah_dataset_training.html')
import csv
from werkzeug.utils import secure_filename

@app.route('/unggah_dataset_training', methods=['GET', 'POST'])
@login_required
def unggah_dataset_training():
    if request.method == 'POST':
        file = request.files['file']
        if not file or (not file.filename.endswith('.csv') and not file.filename.endswith('.xlsx')):
            flash('File harus berformat CSV atau Excel (.xlsx)', 'danger')
            return redirect(request.url)
        filename = secure_filename(file.filename)
        
        # Simpan file ke folder uploads
        UPLOAD_FOLDER = os.path.join(os.getcwd(), "uploads")
        os.makedirs(UPLOAD_FOLDER, exist_ok=True)
        filepath = os.path.join(UPLOAD_FOLDER, filename)
        file.save(filepath)
        conn = get_db()
        try:
            if file.filename.endswith('.xlsx'):
                df = pd.read_excel(filepath)
                # Validasi kolom
                if not all(col in df.columns for col in ['feature1', 'feature2', 'feature3', 'target']):
                    flash('Kolom wajib: feature1, feature2, feature3, target', 'danger')
                    return redirect(request.url)
                # Masukkan baris ke database
                for _, row in df.iterrows():
                    try:
                        bb = float(row['feature1'])
                        tb = float(row['feature2'])
                        lila = float(row['feature3'])
                        status = str(row['target'])
                        conn.execute(
                            'INSERT INTO dataset_training (feature1, feature2, feature3, target) VALUES (?, ?, ?, ?)',
                            (bb, tb, lila, status)
                        )
                    except Exception:
                        continue
            else:
                import csv
                with open(filepath, newline='', encoding='utf-8') as csvfile:
                    reader = csv.DictReader(csvfile)
                    for row in reader:
                        try:
                            bb = float(row['feature1'])
                            tb = float(row['feature2'])
                            lila = float(row['feature3'])
                            status = str(row['target'])
                            conn.execute(
                                'INSERT INTO dataset_training (feature1, feature2, feature3, target) VALUES (?, ?, ?, ?)',
                                (bb, tb, lila, status)
                            )
                        except Exception:
                            continue
            conn.commit()
            flash('Dataset training berhasil diunggah.', 'success')
        finally:
            conn.close()
        return redirect(url_for('kelola_dataset_training'))
    return render_template('dataset/unggah_dataset_training.html')

@app.route('/hapus_semua_dataset_training', methods=['POST'])
@login_required
def hapus_semua_dataset_training():
    conn = get_db()
    conn.execute('DELETE FROM dataset_training')
    conn.commit()
    conn.close()
    flash('Semua data training berhasil dihapus.', 'success')
    return redirect(url_for('kelola_dataset_training'))



# =============================================
# ROUTE LVQ (ADMIN ONLY)
# =============================================

@app.route('/lvq_prototype')
@login_required
def lvq_prototype():
    conn = get_db()
    data_lvq = conn.execute(
        'SELECT id, feature1, feature2, feature3, target FROM dataset_lvq'
    ).fetchall()
    conn.close()
    return render_template('lvq_prototype.html', data_lvq=data_lvq)

@app.route('/lvq_sampling', methods=['POST'])
@login_required
def lvq_sampling():
    conn = get_db()
    data_training = conn.execute('SELECT feature1, feature2, feature3, target FROM dataset_training').fetchall()
    X_train = np.array([[row['feature1'], row['feature2'], row['feature3']] for row in data_training])
    y_train = np.array([row['target'] for row in data_training])
    lvq = LVQ(n_prototypes_per_class=10, learning_rate=0.1, n_epochs=50)
    lvq.fit(X_train, y_train)
    protos, proto_labels = lvq.get_prototypes()
    # Kosongkan tabel prototipe hasil LVQ
    conn.execute('DELETE FROM dataset_lvq')
    # Masukkan prototipe ke dataset_lvq
    for proto, label in zip(protos, proto_labels):
        conn.execute('INSERT INTO dataset_lvq (feature1, feature2, feature3, target) VALUES (?, ?, ?, ?)',
                     (float(proto[0]), float(proto[1]), float(proto[2]), label))
    conn.commit()
    conn.close()
    flash('Proses sampling LVQ selesai. Dataset prototipe siap dipakai KNN!', 'success')
    return redirect(url_for('lvq_prototype'))

@app.route('/lvq_prototype/delete_all', methods=['POST'])
@login_required
def delete_all_lvq_prototype():
    conn = get_db()
    conn.execute('DELETE FROM dataset_lvq')
    conn.commit()
    conn.close()
    flash('Semua prototipe LVQ berhasil dihapus.', 'success')
    return redirect(url_for('lvq_prototype'))

#==============================================
# EVALUASI MODEL 
#==============================================

@app.route('/evaluasi_model')
@login_required
def evaluasi_model():
    conn = get_db()
    try:
        # Ambil parameter KNN
        params = conn.execute('''
            SELECT * FROM parameter_knn 
            ORDER BY created_at DESC LIMIT 1
        ''').fetchone()

        # Ambil statistik LVQ
        lvq_stats = {
            'total_prototypes': conn.execute('SELECT COUNT(*) FROM dataset_lvq').fetchone()[0],
            'total_training': conn.execute('SELECT COUNT(*) FROM dataset_training').fetchone()[0],
            'last_updated': conn.execute('''
                SELECT MAX(created_at) FROM parameter_knn
            ''').fetchone()[0]
        }
        
        if lvq_stats['total_training'] > 0:
            lvq_stats['reduction_ratio'] = (
                (lvq_stats['total_training'] - lvq_stats['total_prototypes']) / 
                lvq_stats['total_training'] * 100
            )
        else:
            lvq_stats['reduction_ratio'] = 0

        # Evaluasi model jika parameter dan dataset tersedia
        model_evaluation = None
        if params and lvq_stats['total_prototypes'] > 0:
            model_evaluation = evaluasi_model_with_parameters(
                params['nilai_k'],
                (params['bobot_berat'], params['bobot_tinggi'], 
                 params['bobot_lila'], params['bobot_umur'], 
                 params['bobot_jk'])
            )

        # Render template dengan semua data yang diperlukan
        return render_template('evaluasi_model.html',
                             params=params,
                             lvq_stats=lvq_stats,
                             model_evaluation=model_evaluation)
                             
    except Exception as e:
        flash(f"Error saat evaluasi model: {str(e)}", "error")
        return render_template('evaluasi_model.html', 
                             params=None,
                             lvq_stats=None,
                             model_evaluation=None)
    finally:
        conn.close()
        
@app.route('/evaluasi_model/data')
@login_required
def get_evaluation_data():
    conn = get_db()
    try:
        params = conn.execute('''
            SELECT * FROM parameter_knn 
            ORDER BY created_at DESC LIMIT 1
        ''').fetchone()
        
        if not params:
            return jsonify({'error': 'Parameter tidak tersedia'})
            
        evaluation = evaluasi_model_with_parameters(
            params['nilai_k'],
            (params['bobot_berat'], params['bobot_tinggi'], 
             params['bobot_lila'], params['bobot_umur'], 
             params['bobot_jk'])
        )
        
        return jsonify({
            'model_evaluation': evaluation,
            'last_updated': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        })
        
    except Exception as e:
        return jsonify({'error': str(e)})
    finally:
        conn.close()

# =============================================
# FUNGSI UPLOAD DATA
# =============================================

@app.route('/balita/upload', methods=['GET', 'POST'])
@login_required
def upload_balita_pengukuran():
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('File tidak ditemukan', 'danger')
            return redirect(url_for('upload_balita_pengukuran'))

        file = request.files['file']
        if file.filename == '':
            flash('Nama file tidak valid', 'danger')
            return redirect(url_for('upload_balita_pengukuran'))

        conn = None
        try:
            df = pd.read_excel(file)
            required_columns = [
                'nik', 'nama', 'nama orangtua', 'jenis kelamin', 'usia_tahun', 'usia_bulan',
                'tgl pengukuran', 'berat badan (KG)', 'Tinggi badan (CM)', 'LILA (CM)'
            ]
            if not all(col in df.columns for col in required_columns):
                flash(f'Kolom wajib: {", ".join(required_columns)}', 'danger')
                return redirect(url_for('upload_balita_pengukuran'))

            conn = get_db()
            notif_pengukuran = []
            for _, row in df.iterrows():
                nik = str(row['nik']).strip()
                nama = row['nama'].strip()
                tahun = int(row['usia_tahun']) if pd.notnull(row['usia_tahun']) else 0
                bulan = int(row['usia_bulan']) if pd.notnull(row['usia_bulan']) else 0

                if pd.notnull(row['tgl pengukuran']):
                    tgl_ukur = pd.to_datetime(row['tgl pengukuran']).date()
                else:
                    tgl_ukur = datetime.today().date()

                tanggal_lahir = tgl_ukur - relativedelta(years=tahun, months=bulan)

                # Cek balita berdasarkan NIK dan nama
                balita = conn.execute(
                    'SELECT id FROM balita WHERE nik = ? AND nama = ?',
                    (nik, nama)
                ).fetchone()

                if not balita:
                    # Tambah balita baru jika belum ada
                    conn.execute(
                        '''INSERT INTO balita (nik, nama, tanggal_lahir, usia_tahun, usia_bulan, jenis_kelamin, nama_ortu)
                           VALUES (?, ?, ?, ?, ?, ?, ?)''',
                        (nik, nama, tanggal_lahir, tahun, bulan, row['jenis kelamin'], row['nama orangtua'])
                    )
                    balita_id = conn.execute('SELECT last_insert_rowid()').fetchone()[0]
                else:
                    # Jika sudah ada, tambahkan notifikasi
                    notif_pengukuran.append('Balita dengan NIK {} dan nama {} sudah ada. Data yang ditambah adalah pengukuran baru.'.format(nik, nama))
                    balita_id = balita['id']

                # Masukkan pengukuran
                berat = float(str(row['berat badan (KG)']).replace(',', '.'))
                tinggi = float(str(row['Tinggi badan (CM)']).replace(',', '.'))
                lila = float(str(row['LILA (CM)']).replace(',', '.'))

                conn.execute(
                    '''INSERT INTO pengukuran (balita_id, tanggal_ukur, berat_badan, tinggi_badan, lingkar_lengan)
                       VALUES (?, ?, ?, ?, ?)''',
                    (
                        balita_id,
                        tgl_ukur,
                        berat,
                        tinggi,
                        lila
                    )
                )
                pengukuran_id = conn.execute('SELECT last_insert_rowid()').fetchone()[0]

                # === KLASIFIKASI OTOMATIS ===
                status_gizi = klasifikasi_knn(berat, tinggi, lila)
                conn.execute(
                    '''INSERT INTO klasifikasi (pengukuran_id, status_gizi, tanggal_klasifikasi)
                       VALUES (?, ?, ?)''',
                    (pengukuran_id, status_gizi, tgl_ukur)
                )
            conn.commit()
            # Batasi jumlah notifikasi agar session cookie tidak terlalu besar!
            max_notif = 5
            if notif_pengukuran:
                for notif in notif_pengukuran[:max_notif]:
                    flash(notif, 'info')
                if len(notif_pengukuran) > max_notif:
                    flash(f"dan {len(notif_pengukuran)-max_notif} balita lainnya sudah ada...", "info")
            flash('Data balita, pengukuran, dan status gizi berhasil diunggah', 'success')
        except Exception as e:
            flash(f'Error: {str(e)}', 'danger')
        finally:
            if conn is not None:
                conn.close()
        return redirect(url_for('upload_balita_pengukuran'))
    return render_template('balita/upload.html')

# =============================================
# FUNGSI PENDUKUNG
# =============================================

def get_knn_parameters():
    """Mengambil parameter KNN dari database"""
    conn = get_db()
    params = conn.execute('SELECT * FROM parameter_knn').fetchone()
    conn.close()
    return params

def klasifikasi_knn(berat_badan, tinggi_badan, lingkar_lengan):
    conn = get_db()
    data_protos = conn.execute('SELECT feature1, feature2, feature3, target FROM dataset_lvq').fetchall()
    if not data_protos:
        return "Dataset LVQ kosong, lakukan sampling LVQ dahulu"
        
    X_train = [[row['feature1'], row['feature2'], row['feature3']] for row in data_protos]
    y_train = [row['target'] for row in data_protos]
    
    params = conn.execute('SELECT * FROM parameter_knn').fetchone()
    nilai_k = params['nilai_k'] if params else 3
    bobot = [
        float(params['bobot_berat']) if params else 1,
        float(params['bobot_tinggi']) if params else 1,
        float(params['bobot_lila']) if params else 1
    ]
    
    model = KNN(k=nilai_k, bobot=bobot)
    model.fit(X_train, y_train)
    fitur_balita = [[berat_badan, tinggi_badan, lingkar_lengan]]
    status_gizi = model.predict(fitur_balita)[0]
    return status_gizi

# =============================================
# JALANKAN APLIKASI
# =============================================

print("Daftar file di folder templates:", os.listdir(os.path.join(app.root_path, 'templates')))
if __name__ == '__main__':
    app.run(debug=True)